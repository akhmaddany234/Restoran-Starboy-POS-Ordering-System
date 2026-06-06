from tkinter import *
from tkinter.messagebox import *
from PIL import ImageTk, Image
import openpyxl
from datetime import datetime
import os
from tkinter import ttk
import locale

class PemilikMixin:
#-----------------------------------------------------------------------------KODE UNTUK LOGIN PEMILIK-----------------------------------------------------------------------------------------------------------------------
    def loginpemilik(self):
        self.tampilan.destroy()
        self.window_pemilik = Tk()
        self.window_pemilik.title("Pemilik")
        self.window_pemilik.geometry('1280x680')
        
        # Memasukkan gambar tampilan login akun pemilik
        img1 = Image.open("tampilan/loginpemilik.jpg")
        self.pemiliklogin = ImageTk.PhotoImage(img1)
        label_gambarpemilik = Label(self.window_pemilik, image=self.pemiliklogin)
        label_gambarpemilik.place(y=0)

        # Pemasukan email
        Email_ = Label(self.window_pemilik, text="EMAIL:", font= ("Times New Roman",25,'bold'),bg='#FCF6DB',width=10,height=0)
        Email_.place(x=585-40, y=210-20)
        self.emailpemilik = StringVar()
        MasukkanEmail = Entry(self.window_pemilik,textvariable=self.emailpemilik,font=("Times New Roman",20),borderwidth=0,width=20,justify="center") #justify dibuat untuk tulisan dari tengah
        MasukkanEmail.place(x=542-40,y=263-16)

        # Pemasukan sandi
        Sandi_ = Label(self.window_pemilik, text="SANDI:", font= ("Times New Roman",25,'bold'),bg='#FCF6DB',width=10,height=0)
        Sandi_.place(x=585-40, y=328-25)
        self.sandipemilik = StringVar()
        MasukkanSandi = Entry(self.window_pemilik,textvariable=self.sandipemilik,font=("Times New Roman",20),borderwidth=0,width=20,justify="center",show="*") #justify dibuat untuk tulisan dari tengah
        MasukkanSandi.place(x=542-40,y=380-24)

        # Tombol button login
        tombol_login = Button(self.window_pemilik,text="LOGIN", font=("Times New Roman",20,"bold"),command=self.cekakunpemilik,bg="#7D0B00",fg="white",width=10,cursor='hand2')
        tombol_login.place(x=560,y=420)

        # Cek akun pemilik apakah true or false
    def cekakunpemilik(self):
        Email = self.emailpemilik.get()
        Sandi = self.sandipemilik.get()

        if Email == "dany" and Sandi == "dany123":
            showinfo("berhasil","kamu berhasil login")
            self.masukakunpemilik()
        else:
            showerror("Gagal", "mungkin email atau sandi anda salah, silahkan cobalagi!")

#-----------------------------------------------------------------------------KODE HISTORI PENJUALAN--------------------------------------------------------------------------------------------------------------
    def masukakunpemilik(self):
        self.window_pemilik.destroy()
        self.window_akunpemilik = Tk()
        self.window_akunpemilik.geometry('1280x680')
        
        img2 = Image.open("tampilan/11.jpg")  
        self.pemasukan = ImageTk.PhotoImage(img2)
        label_gambar_akun_pemilik = Label(self.window_akunpemilik, image=self.pemasukan)
        label_gambar_akun_pemilik.place(y=0)

        Button(self.window_akunpemilik,text="HARIAN",bg="#7d0b00",fg="white",bd=0,activebackground="#7d0b00",font=('times new roman',15,'bold'),width=23,cursor='hand2',command=self.rekap_harian).place(x=510,y=349)
        Button(self.window_akunpemilik,text="   BULANAN   ",bg="#7d0b00",fg="white",bd=0,activebackground="#7d0b00",font=('times new roman',15,'bold'),width=23,cursor='hand2',command=self.RekapBulanan).place(x=510,y=432)

    def rekap_harian(self):
        self.window_akunpemilik.withdraw()
        self.window_RiwayatHarian = Toplevel(self.window_akunpemilik)
        self.window_RiwayatHarian.geometry('1280x680')

        # Memuat dan menampilkan gambar latar belakang
        img14 = Image.open("tampilan/LAPORAN HARIAN.png")
        self.FotoRiwayatHarian = ImageTk.PhotoImage(img14)
        label_gambar = Label(self.window_RiwayatHarian, image=self.FotoRiwayatHarian)
        label_gambar.pack(pady=0)

        Kembali = Button(self.window_RiwayatHarian,text='KEMBALI',bg='#7d0b00', fg="white",borderwidth=0,activebackground='#7d0b00',font=('times new roman',18,'bold'),width=22,cursor='hand2',command=self.Destroy)
        Kembali.place(x=480, y=615)

        if os.path.exists("data_pelanggan/last_run.txt"):  #
            with open("data_pelanggan/last_run.txt", "r") as file:
                last_run_date_str = file.read().strip()
                last_run_date = datetime.strptime(last_run_date_str, "%Y-%m-%d").date()

            today = datetime.now().date()

            # Membandingkan tanggal sekarang dengan tanggal terakhir
            if today > last_run_date:
                self.workbook = openpyxl.load_workbook('data_pelanggan/REKAP_HARIAN.xlsx')
                self.sheet = self.workbook.active

                self.sheet['A1'] = 0
                self.sheet['A2'] = 0
                self.sheet['A3'] = 0
                self.sheet['A4'] = 0
                self.sheet['A5'] = 0
                self.sheet['A6'] = 0
                self.sheet['A7'] = 0
                self.sheet['A8'] = 0
                self.sheet['A9'] = 0

                self.sheet['B1'] = 0
                self.sheet['B2'] = 0
                self.sheet['B3'] = 0
                self.sheet['B4'] = 0
                self.sheet['B5'] = 0
                self.sheet['B6'] = 0
                self.sheet['B7'] = 0
                self.sheet['B8'] = 0

                self.sheet['C1'] = 0
                self.sheet['C2'] = 0
                self.sheet['C3'] = 0
                self.sheet['C4'] = 0
                self.sheet['C5'] = 0
                self.sheet['C6'] = 0
                self.sheet['C7'] = 0
                self.sheet['C8'] = 0

                self.workbook.save('data_pelanggan/REKAP_HARIAN.xlsx')

        # Menyimpan tanggal terakhir kali program dijalankan
        with open("data_pelanggan/last_run.txt", "w") as file:
            file.write(datetime.now().date().strftime("%Y-%m-%d"))

        self.workbook = openpyxl.load_workbook('REKAP HARIAN.xlsx') #load_workbook untuk membuka pyxl yang  sudah ada
        self.sheet = self.workbook.active

        # Makanan
        Label(self.window_RiwayatHarian,text=self.sheet['A1'].value,bg='white',font=('Times new roman',18,'bold')).place(x=400,y=207)
        Label(self.window_RiwayatHarian,text=self.sheet['A2'].value,bg='white',font=('Times new roman',18,'bold')).place(x=400,y=235)
        Label(self.window_RiwayatHarian,text=self.sheet['A3'].value,bg='white',font=('Times new roman',18,'bold')).place(x=400,y=235+28)
        Label(self.window_RiwayatHarian,text=self.sheet['A4'].value,bg='white',font=('Times new roman',18,'bold')).place(x=400,y=235+28*2)
        Label(self.window_RiwayatHarian,text=self.sheet['A5'].value,bg='white',font=('Times new roman',18,'bold')).place(x=400,y=235+28*3)
        Label(self.window_RiwayatHarian,text=self.sheet['A6'].value,bg='white',font=('Times new roman',18,'bold')).place(x=400,y=235+28*4)
        Label(self.window_RiwayatHarian,text=self.sheet['A7'].value,bg='white',font=('Times new roman',18,'bold')).place(x=400,y=235+28*5)
        Label(self.window_RiwayatHarian,text=self.sheet['A8'].value,bg='white',font=('Times new roman',18,'bold')).place(x=400,y=235+28*6)
        Label(self.window_RiwayatHarian,text=self.sheet['A9'].value,bg='white',font=('Times new roman',18,'bold')).place(x=400,y=235+28*7)

        # Minuman
        Label(self.window_RiwayatHarian,text=self.sheet['B1'].value,bg='white',font=('Times new roman',18,'bold')).place(x=800,y=207)
        Label(self.window_RiwayatHarian,text=self.sheet['B2'].value,bg='white',font=('Times new roman',18,'bold')).place(x=800,y=235)
        Label(self.window_RiwayatHarian,text=self.sheet['B3'].value,bg='white',font=('Times new roman',18,'bold')).place(x=800,y=235+28*1)
        Label(self.window_RiwayatHarian,text=self.sheet['B4'].value,bg='white',font=('Times new roman',18,'bold')).place(x=800,y=235+28*2)
        Label(self.window_RiwayatHarian,text=self.sheet['B5'].value,bg='white',font=('Times new roman',18,'bold')).place(x=800,y=235+28*3)
        Label(self.window_RiwayatHarian,text=self.sheet['B6'].value,bg='white',font=('Times new roman',18,'bold')).place(x=800,y=235+28*4)
        Label(self.window_RiwayatHarian,text=self.sheet['B7'].value,bg='white',font=('Times new roman',18,'bold')).place(x=800,y=235+28*5)
        Label(self.window_RiwayatHarian,text=self.sheet['B8'].value,bg='white',font=('Times new roman',18,'bold')).place(x=800,y=235+28*6)

        # Jajan
        Label(self.window_RiwayatHarian,text=self.sheet['C1'].value,bg='white',font=('Times new roman',18,'bold')).place(x=1150,y=207)
        Label(self.window_RiwayatHarian,text=self.sheet['C2'].value,bg='white',font=('Times new roman',18,'bold')).place(x=1150,y=235)
        Label(self.window_RiwayatHarian,text=self.sheet['C3'].value,bg='white',font=('Times new roman',18,'bold')).place(x=1150,y=235+28*1)
        Label(self.window_RiwayatHarian,text=self.sheet['C4'].value,bg='white',font=('Times new roman',18,'bold')).place(x=1150,y=235+28*2)
        Label(self.window_RiwayatHarian,text=self.sheet['C5'].value,bg='white',font=('Times new roman',18,'bold')).place(x=1150,y=235+28*3)
        Label(self.window_RiwayatHarian,text=self.sheet['C6'].value,bg='white',font=('Times new roman',18,'bold')).place(x=1150,y=235+28*4)
        Label(self.window_RiwayatHarian,text=self.sheet['C7'].value,bg='white',font=('Times new roman',18,'bold')).place(x=1150,y=235+28*5)
        Label(self.window_RiwayatHarian,text=self.sheet['C8'].value,bg='white',font=('Times new roman',18,'bold')).place(x=1150,y=235+28*6)

        self.current_day = self.get_current_day()

        # Membuat label untuk menampilkan nama bulan
        label_month = ttk.Label(self.window_RiwayatHarian, text=self.current_day,background='#ffb4ae',font=('times new roman',20,'bold'))
        label_month.place(x=690, y=93)
    def get_current_day(self):
        locale.setlocale(locale.LC_TIME, 'id_ID')
        self.now = datetime.now()
        return self.now.strftime("%A, %d %B %Y").title()
    def Destroy(self):
        self.window_RiwayatHarian.destroy()
        self.window_akunpemilik.deiconify()

